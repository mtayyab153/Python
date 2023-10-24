import openpyxl
from openpyxl import Workbook

# wb = Workbook()
#
# ws = wb.active

# ws1 = wb.create_sheet("sheet1", 0)
# ws2 = wb.create_sheet("sheet2")
# ws3 = wb.create_sheet("sheet3", -1)

# ws["A1"] = "cell 1"
# ws["B2"] = "cell 2"

# for i in range(1,11):
#     ws[f"A{i}"] = i

# for i in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
#     for j in range(1,11):
#         ws[f"{i}{j}"] = f"{i}{j}"


# wb.save(filename="example.xlsx")


# wb = openpyxl.load_workbook("example.xlsx")
#
# ws = wb.active
#
# # print(ws["A1"].value)
#
# for i in range(1,6):
#     for j in range(1,6):
#         print(i, j, ws.cell(row=i, column=j).value)

# wb = openpyxl.load_workbook("example2.xlsx")
#
# ws = wb.active
#
# ws["A11"] = "=SUM(A1:A10)"

# wb.save("example2.xlsx")